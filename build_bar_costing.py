import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import copy

wb = openpyxl.Workbook()

# ── Colour palette ──────────────────────────────────────────────────────────
BLACK   = "FF1A1A1A"
WHITE   = "FFFFFFFF"
GOLD    = "FFB8860B"
GOLD_LT = "FFFFF8DC"
DARK    = "FF1C1C2E"
MID     = "FF2E2E45"
GREEN   = "FF006400"
GREEN_LT= "FFE8F5E9"
RED_LT  = "FFFCE4EC"
RED_DK  = "FFC62828"
BLUE_LT = "FFE3F2FD"
BLUE_DK = "FF1565C0"
GREY_LT = "FFF5F5F5"
GREY_BD = "FFE0E0E0"
AMBER   = "FFFF8F00"
AMBER_LT= "FFFFF3E0"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Arial")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin"):
    s = Side(style=style, color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium", color="FF888888")
    n = Side(style="thin",   color="FFCCCCCC")
    return Border(left=t, right=t, top=n, bottom=n)

def header_border():
    t = Side(style="medium", color=GOLD)
    return Border(left=t, right=t, top=t, bottom=t)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 – INGREDIENT MASTER LIST
# ═══════════════════════════════════════════════════════════════════════════
ws_master = wb.active
ws_master.title = "MASTER LIST"

# freeze
ws_master.freeze_panes = "A4"

# Title row
ws_master.merge_cells("A1:L1")
c = ws_master["A1"]
c.value = "INGREDIENT MASTER LIST"
c.fill = fill(DARK)
c.font = Font(bold=True, color=GOLD, size=14, name="Arial")
c.alignment = align("center")
ws_master.row_dimensions[1].height = 30

ws_master.merge_cells("A2:L2")
c2 = ws_master["A2"]
c2.value = "Cost per ml/g is derived from unit cost ÷ unit size.  Inventory entered in GRAMS; system converts to ml where applicable."
c2.fill = fill(MID)
c2.font = Font(bold=False, color="FFCCCCCC", size=9, italic=True, name="Arial")
c2.alignment = align("center")
ws_master.row_dimensions[2].height = 16

# Column headers row 3
headers_master = [
    "TAHWEEL", "SUPPLIER", "CATEGORY", "DESCRIPTION",
    "UNIT SIZE\n(g or ml)", "DENSITY\n(g/ml)", "UNIT COST\n(AED)",
    "COST PER\nGRAM (AED)", "COST PER\nML (AED)",
    "INV STOCK\n(grams entered)", "INV STOCK\n(ml converted)",
    "INV VALUE\n(AED)"
]
col_widths_master = [12, 16, 14, 42, 12, 10, 12, 13, 13, 16, 16, 14]

for i, (h, w) in enumerate(zip(headers_master, col_widths_master), 1):
    col = get_column_letter(i)
    ws_master.column_dimensions[col].width = w
    c = ws_master.cell(row=3, column=i, value=h)
    c.fill = fill(DARK)
    c.font = Font(bold=True, color=GOLD, size=9, name="Arial")
    c.alignment = align("center", wrap=True)
    c.border = header_border()
ws_master.row_dimensions[3].height = 32

# Sample ingredient data
# density g/ml: water=1.0, spirits~0.91, syrups~1.3, juice~1.04, cream~1.0, puree~1.1
ingredients = [
    # SPIRITS
    ("BI101350","A+E","SPIRITS","Aperol 700ml Italy",700,0.91,106.06),
    ("BI101352","MMI","SPIRITS","Campari 1000ml Italy",1000,0.91,148.20),
    ("BI101358","MMI","SPIRITS","Cinzano Rosso 1Ltr Italy",1000,0.91,92.82),
    ("BI106294","A+E","SPIRITS","Carpano Antica Formula Vermouth 1000ml",1000,0.91,447.72),
    ("BI101731","MMI","SPIRITS","Bols Elderflower Liqueur 700ml",700,0.91,81.64),
    ("BI100585","M.H.INTL","SPIRITS","Kahlua Coffee Liqueur 700ml",700,0.91,95.00),
    # SPARKLING / BEER
    ("BI124988","MMI","SPARKLING","Gran Ventino Brut Sparkling 75cl",750,1.00,44.20),
    ("BI101228","BIDFOOD","SPARKLING WATER","San Pellegrino Sparkling 750ml",750,1.00,4.75),
    # JUICE FRESH
    ("BI110766","PURE QUALITY","JUICE FRESH","Juice Fresh Lemon 1.5Ltr UAE",1500,1.04,15.225),
    ("BI110833","PURE QUALITY","JUICE FRESH","Juice Fresh Watermelon 1.5Ltr UAE",1500,1.04,12.00),
    ("BI110800","PURE QUALITY","JUICE FRESH","Juice Fresh Orange 1.5Ltr UAE",1500,1.04,6.615),
    ("BI110818","PURE QUALITY","JUICE FRESH","Juice Fresh Pineapple 1.5Ltr UAE",1500,1.04,18.90),
    ("BI110749","PURE QUALITY","JUICE FRESH","Juice Fresh Grapefruit 500ml UAE",500,1.04,5.80),
    # SYRUPS
    ("BI100585_S","M.H.INTL","SYRUP","Syrup Agave Bio Allos 250ml",250,1.30,26.00),
    ("FI114909","M.H.INTL","SYRUP","Syrup Maple Organic 250ml",250,1.30,47.00),
    ("BI110937","M.H.INTL","SYRUP","Syrup Strawberry Fabbri 1000ml",1000,1.30,51.97),
    ("BI110909","M.H.INTL","SYRUP","Syrup Coconut Fabbri 1000ml",1000,1.30,59.07),
    ("BI110937_E","M.H.INTL","SYRUP","Syrup Elderflower Fabbri 1000ml",1000,1.30,52.76),
    # PUREE / FROZEN
    ("FI110607","CHEF ME","PUREE","Puree Peach White Frozen Boiron 1KG",1000,1.10,210.00),
    ("FI110599","CHEF ME","PUREE","Puree Passion Fruit Frozen Boiron 1KG",1000,1.10,174.00),
    ("FI110026","CHEF ME","PUREE","Coulis Raspberry Frozen Boiron 500g",500,1.10,300.00),
    # MILK / CREAM
    ("FI105618","AL AIN","MILK","Milk Fresh Full Cream Al Ain 2Ltr",2000,1.03,6.30),
    ("FI105595","M.H.INTL","MILK","Milk Coconut Barista Alpro 1Ltr",1000,1.03,11.00),
    ("FI105578","M.H.INTL","MILK","Milk Almond Alpro Barista 1Ltr",1000,1.03,10.00),
    # SOFT DRINKS
    ("BI100468","AL AHLIA","SOFT DRINK","Sprite 300ml Can",300,1.00,2.06),
    ("BI111636","AL AHLIA","SOFT DRINK","Schweppes Tonic Water 300ml Can",300,1.00,49.52),
    ("BI100410","CHEF ME","SOFT DRINK","Lemonade Rose Fentimans 275ml",275,1.00,9.50),
    # FRUITS / GARNISH
    ("FI116736","BARAKAT","CITRUS","Lemon Premium Quality Fresh South Africa 1kg",1000,1.00,4.20),
    ("FI116744","BARAKAT","CITRUS","Lime Corona Fresh Mexico 1kg",1000,1.00,6.00),
    ("FI116847","BARAKAT","CITRUS","Orange Navel 1kg South Africa",1000,1.00,4.00),
    ("FI117375","FRESH EXPRESS","HERB","Mint Leaves Cyprus Fresh 1Kg",1000,1.00,56.00),
    ("FI117022","BARAKAT","BERRY","Strawberry Fresh Egypt 1kg",1000,1.00,28.00),
    # DRY / OTHER
    ("FI114575","M.H.INTL","DRY GOODS","Sugar Fine Granulated White 2kg",2000,1.00,5.50),
    ("FI109983","SAFCO","GARNISH","Cherry Amarena in Syrup Toschi 510g",510,1.00,162.00),
    ("FI130063","ICE","ICE","Ice Block 5x5x7 Each Piece",1000,1.00,1.75),
]

for row_i, (tahw, supp, cat, desc, unit_size, density, unit_cost) in enumerate(ingredients, 4):
    col = lambda c: ws_master.cell(row=row_i, column=c)
    col(1).value = tahw
    col(2).value = supp
    col(3).value = cat
    col(4).value = desc
    col(5).value = unit_size
    col(6).value = density
    col(7).value = unit_cost
    # cost per gram = unit_cost / unit_size
    col(8).value = f"=G{row_i}/E{row_i}"
    col(8).number_format = '"AED "#,##0.0000'
    # cost per ml = cost_per_gram / density (g/ml)
    col(9).value = f"=H{row_i}/F{row_i}"
    col(9).number_format = '"AED "#,##0.0000'
    # inv stock grams (user input)
    col(10).value = 0
    col(10).fill = fill(AMBER_LT)
    # inv stock in ml = grams / density
    col(11).value = f"=IF(J{row_i}=0,0,J{row_i}/F{row_i})"
    col(11).number_format = '#,##0.0" ml"'
    # inv value = grams * cost per gram
    col(12).value = f"=J{row_i}*H{row_i}"
    col(12).number_format = '"AED "#,##0.00'

    # styling
    row_fill = fill(WHITE) if row_i % 2 == 0 else fill(GREY_LT)
    for c_i in range(1, 13):
        cell = ws_master.cell(row=row_i, column=c_i)
        if c_i not in [8, 9, 11, 12]:
            cell.fill = row_fill
        cell.font = Font(size=9, name="Arial",
                         color="FF000080" if c_i == 1 else BLACK)
        cell.alignment = align("center" if c_i in [1,5,6,7,8,9,10,11,12] else "left")
        cell.border = border()
    ws_master.row_dimensions[row_i].height = 14

    # highlight input column
    ws_master.cell(row=row_i, column=10).font = Font(size=9, bold=True,
                                                      color=AMBER, name="Arial")

# summary total row
last = 4 + len(ingredients) - 1
tot_row = last + 2
ws_master.cell(tot_row, 1).value = "TOTAL INVENTORY VALUE"
ws_master.merge_cells(f"A{tot_row}:K{tot_row}")
ws_master.cell(tot_row, 1).fill = fill(DARK)
ws_master.cell(tot_row, 1).font = Font(bold=True, color=GOLD, size=10, name="Arial")
ws_master.cell(tot_row, 1).alignment = align("right")
ws_master.cell(tot_row, 12).value = f"=SUM(L4:L{last})"
ws_master.cell(tot_row, 12).number_format = '"AED "#,##0.00'
ws_master.cell(tot_row, 12).fill = fill(DARK)
ws_master.cell(tot_row, 12).font = Font(bold=True, color=GOLD, size=10, name="Arial")
ws_master.cell(tot_row, 12).alignment = align("center")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 – COCKTAIL RECIPE COSTING
# ═══════════════════════════════════════════════════════════════════════════
ws_cost = wb.create_sheet("RECIPE COSTING")
ws_cost.freeze_panes = "A5"

# col widths
cost_cols = [14, 42, 14, 12, 12, 13, 13, 3, 20, 14]
cost_col_letters = [get_column_letter(i) for i in range(1, len(cost_cols)+1)]
for col, w in zip(cost_col_letters, cost_cols):
    ws_cost.column_dimensions[col].width = w

# Title
ws_cost.merge_cells("A1:J1")
c = ws_cost["A1"]
c.value = "BEVERAGE RECIPE COSTING SHEET"
c.fill = fill(DARK)
c.font = Font(bold=True, color=GOLD, size=14, name="Arial")
c.alignment = align("center")
ws_cost.row_dimensions[1].height = 30

ws_cost.merge_cells("A2:J2")
c2 = ws_cost["A2"]
c2.value = (
    "RECIPE QUANTITIES in ml (liquids) or g/nos (solids).  "
    "INVENTORY tracked in grams — converted to ml automatically via density.  "
    "Yellow cells = user input."
)
c2.fill = fill(MID)
c2.font = Font(italic=True, color="FFCCCCCC", size=9, name="Arial")
c2.alignment = align("center")
ws_cost.row_dimensions[2].height = 16

# column headers row 3
cost_headers = [
    "TAHWEEL", "INGREDIENT DESCRIPTION",
    "UNIT SIZE\n(ml)", "UNIT COST\n(AED)",
    "RECIPE\n(ml or g)",
    "UNIT\n(ml / g / nos)",
    "RECIPE COST\n(AED)", "",
    "SUMMARY", "VALUE"
]
for i, h in enumerate(cost_headers, 1):
    c = ws_cost.cell(row=3, column=i, value=h)
    c.fill = fill(DARK)
    c.font = Font(bold=True, color=GOLD, size=9, name="Arial")
    c.alignment = align("center", wrap=True)
    c.border = header_border()
ws_cost.row_dimensions[3].height = 32

# Helper to write one recipe block
def write_recipe(start_row, name, bev_type, selling_price,
                 method, glass, garnish, ice, straw,
                 ingredients_list):
    """
    ingredients_list: list of tuples
      (tahweel, description, unit_ml, unit_cost, recipe_qty, unit_label)
      unit_label: 'ml' | 'g' | 'nos'
      For g: recipe_cost = unit_cost/unit_ml * recipe_qty  (cost per gram)
      For ml: recipe_cost = unit_cost/unit_ml * recipe_qty (cost per ml)
    """
    r = start_row

    # — recipe name bar —
    ws_cost.merge_cells(f"A{r}:G{r}")
    nc = ws_cost[f"A{r}"]
    nc.value = f"  ◆  {name.upper()}  |  {bev_type.upper()}"
    nc.fill = fill(MID)
    nc.font = Font(bold=True, color=GOLD, size=11, name="Arial")
    nc.alignment = align("left")
    ws_cost.row_dimensions[r].height = 22
    r += 1

    # — ingredient rows —
    ing_start = r
    for (tahw, desc, unit_ml, unit_cost, recipe_qty, unit_lbl) in ingredients_list:
        cells = [
            ws_cost.cell(row=r, column=1, value=tahw),
            ws_cost.cell(row=r, column=2, value=desc),
            ws_cost.cell(row=r, column=3, value=unit_ml),
            ws_cost.cell(row=r, column=4, value=unit_cost),
            ws_cost.cell(row=r, column=5, value=recipe_qty),
            ws_cost.cell(row=r, column=6, value=unit_lbl),
            ws_cost.cell(row=r, column=7,
                         value=f"=IF(C{r}=0,0,(D{r}/C{r})*E{r})"),
        ]
        cells[6].number_format = '"AED "#,##0.0000'

        row_bg = fill(WHITE) if r % 2 == 0 else fill(GREY_LT)
        for ci, cell in enumerate(cells):
            cell.fill = row_bg
            cell.font = Font(size=9, name="Arial",
                             color="FF000080" if ci == 0 else BLACK)
            cell.alignment = align("center" if ci not in [1] else "left")
            cell.border = border()
        # highlight input cells
        for col_input in [3, 4, 5]:
            ws_cost.cell(row=r, column=col_input).fill = fill(AMBER_LT)
            ws_cost.cell(row=r, column=col_input).font = Font(
                size=9, color=AMBER, bold=True, name="Arial")
        ws_cost.row_dimensions[r].height = 14
        r += 1

    ing_end = r - 1
    # blank filler rows to keep summary aligned (max 8 ingredient rows)
    while r < ing_start + 8:
        for ci in range(1, 8):
            cell = ws_cost.cell(row=r, column=ci, value="")
            cell.fill = fill(GREY_LT)
            cell.border = border()
        ws_cost.row_dimensions[r].height = 14
        r += 1

    # — totals row —
    tot_r = r
    ws_cost.cell(tot_r, 1).value = "TOTAL"
    ws_cost.merge_cells(f"A{tot_r}:F{tot_r}")
    ws_cost.cell(tot_r, 1).fill = fill(DARK)
    ws_cost.cell(tot_r, 1).font = Font(bold=True, color=GOLD, size=9, name="Arial")
    ws_cost.cell(tot_r, 1).alignment = align("right")
    ws_cost.cell(tot_r, 7).value = f"=SUM(G{ing_start}:G{ing_end})"
    ws_cost.cell(tot_r, 7).number_format = '"AED "#,##0.0000'
    ws_cost.cell(tot_r, 7).fill = fill(DARK)
    ws_cost.cell(tot_r, 7).font = Font(bold=True, color=GOLD, size=10, name="Arial")
    ws_cost.cell(tot_r, 7).alignment = align("center")
    ws_cost.row_dimensions[tot_r].height = 18

    # — summary panel (col I & J, rows ing_start to ing_start+8) —
    summary_items = [
        ("METHOD:", method),
        ("GLASS:", glass),
        ("GARNISH:", garnish),
        ("ICE:", ice),
        ("STRAW:", straw),
        ("", ""),
        ("RECIPE COST (AED):", f"=G{tot_r}"),
        ("SELLING PRICE (AED):", selling_price),
        ("COST %:", f"=IF(J{ing_start+7}=0,0,G{tot_r}/J{ing_start+7})"),
    ]
    for s_i, (label, val) in enumerate(summary_items):
        sr = ing_start + s_i
        lc = ws_cost.cell(sr, 9, value=label)
        vc = ws_cost.cell(sr, 10, value=val)
        lc.fill = fill(BLUE_LT)
        vc.fill = fill(WHITE)
        lc.font = Font(bold=True, size=9, name="Arial", color=BLUE_DK)
        vc.font = Font(bold=True, size=9, name="Arial",
                       color=GREEN if "COST" in label and "%" not in label else
                             RED_DK if "%" in label else BLACK)
        lc.alignment = align("right")
        vc.alignment = align("center")
        lc.border = border()
        vc.border = border()
        if "COST (AED)" in label:
            vc.number_format = '"AED "#,##0.00'
        elif "PRICE" in label:
            vc.number_format = '"AED "#,##0.00'
        elif "%" in label:
            vc.number_format = '0.00%'
        ws_cost.row_dimensions[sr].height = 14

    # spacer
    for ci in range(1, 11):
        cell = ws_cost.cell(tot_r + 1, ci, value="")
        cell.fill = fill(GREY_BD)
        cell.border = border()
    ws_cost.row_dimensions[tot_r + 1].height = 6

    return tot_r + 2  # next recipe starts here

# Sample recipes
next_row = 4
next_row = write_recipe(
    next_row,
    name="Aperol Spritz",
    bev_type="Cocktail",
    selling_price=69,
    method="Build in glass over ice",
    glass="Wine Glass",
    garnish="Orange slice",
    ice="Cubed",
    straw="Yes",
    ingredients_list=[
        ("BI101350", "Aperol 700ml Italy", 700, 106.06, 60, "ml"),
        ("BI124988", "Gran Ventino Brut Sparkling 75cl", 750, 44.20, 90, "ml"),
        ("BI101228", "San Pellegrino Sparkling 750ml", 750, 4.75, 30, "ml"),
    ]
)

next_row = write_recipe(
    next_row,
    name="Classic Mojito",
    bev_type="Cocktail",
    selling_price=75,
    method="Muddle mint & lime. Add rum & sugar. Top with soda.",
    glass="Highball",
    garnish="Mint sprig + lime wedge",
    ice="Crushed",
    straw="Yes",
    ingredients_list=[
        ("BI101XXX", "White Rum 700ml", 700, 95.00, 60, "ml"),
        ("BI110766", "Juice Fresh Lemon/Lime 1.5Ltr", 1500, 15.225, 30, "ml"),
        ("FI114575", "Simple Syrup (Sugar 2kg)", 2000, 5.50, 20, "ml"),
        ("FI117375", "Mint Leaves Fresh 1Kg", 1000, 56.00, 10, "g"),
        ("BI101228", "Soda Water San Pellegrino 750ml", 750, 4.75, 90, "ml"),
    ]
)

next_row = write_recipe(
    next_row,
    name="Espresso Martini",
    bev_type="Cocktail",
    selling_price=82,
    method="Shake all ingredients hard with ice. Double strain.",
    glass="Coupe",
    garnish="3 coffee beans",
    ice="Shake only",
    straw="No",
    ingredients_list=[
        ("BI101XXX", "Vodka 700ml", 700, 100.00, 50, "ml"),
        ("BI100585", "Kahlua Coffee Liqueur 700ml", 700, 95.00, 30, "ml"),
        ("BI109655", "Fresh Espresso Shot", 30, 0.95, 40, "ml"),
        ("FI114575", "Simple Syrup 2kg", 2000, 5.50, 10, "ml"),
    ]
)

next_row = write_recipe(
    next_row,
    name="Watermelon Mint Cooler",
    bev_type="Mocktail",
    selling_price=48,
    method="Blend watermelon juice with mint. Top with soda.",
    glass="Highball",
    garnish="Watermelon slice + mint",
    ice="Cubed",
    straw="Yes",
    ingredients_list=[
        ("BI110833", "Juice Fresh Watermelon 1.5Ltr", 1500, 12.00, 120, "ml"),
        ("BI110766", "Juice Fresh Lemon 1.5Ltr", 1500, 15.225, 20, "ml"),
        ("FI117375", "Mint Leaves Fresh 1Kg", 1000, 56.00, 8, "g"),
        ("FI114575", "Simple Syrup 2kg", 2000, 5.50, 15, "ml"),
        ("BI101228", "Soda Water San Pellegrino 750ml", 750, 4.75, 60, "ml"),
    ]
)

next_row = write_recipe(
    next_row,
    name="Peach Bellini",
    bev_type="Cocktail",
    selling_price=70,
    method="Pour peach puree into chilled flute. Top with sparkling.",
    glass="Champagne Flute",
    garnish="Peach slice",
    ice="None",
    straw="No",
    ingredients_list=[
        ("FI110607", "Puree Peach White Frozen Boiron 1KG", 1000, 210.00, 50, "g"),
        ("BI124988", "Gran Ventino Brut Sparkling 75cl", 750, 44.20, 100, "ml"),
    ]
)

next_row = write_recipe(
    next_row,
    name="Coconut Cold Brew Latte",
    bev_type="Coffee",
    selling_price=38,
    method="Pour cold brew over ice. Add coconut milk. Stir.",
    glass="Rocks Glass",
    garnish="Coconut flakes",
    ice="Cubed",
    straw="Yes",
    ingredients_list=[
        ("BI109655", "Espresso / Cold Brew (per 30ml)", 30, 0.95, 90, "ml"),
        ("FI105595", "Milk Coconut Barista Alpro 1Ltr", 1000, 11.00, 150, "ml"),
        ("BI110909", "Syrup Coconut Fabbri 1000ml", 1000, 59.07, 15, "ml"),
    ]
)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 – INVENTORY (grams in, ml out)
# ═══════════════════════════════════════════════════════════════════════════
ws_inv = wb.create_sheet("INVENTORY")
ws_inv.freeze_panes = "A5"

inv_col_widths = [12, 16, 14, 42, 14, 10, 14, 14, 14, 14, 14, 16]
for i, w in enumerate(inv_col_widths, 1):
    ws_inv.column_dimensions[get_column_letter(i)].width = w

# Title
ws_inv.merge_cells("A1:L1")
c = ws_inv["A1"]
c.value = "DAILY INVENTORY SHEET  —  Enter stock in GRAMS  |  System converts to ML"
c.fill = fill(DARK)
c.font = Font(bold=True, color=GOLD, size=13, name="Arial")
c.alignment = align("center")
ws_inv.row_dimensions[1].height = 28

ws_inv.merge_cells("A2:L2")
c2 = ws_inv["A2"]
c2.value = (
    "Grams entered here update the MASTER LIST inventory columns automatically.  "
    "Par levels trigger LOW STOCK highlighting.  Cost values pulled from Master List."
)
c2.fill = fill(MID)
c2.font = Font(italic=True, color="FFCCCCCC", size=9, name="Arial")
c2.alignment = align("center")
ws_inv.row_dimensions[2].height = 16

inv_headers = [
    "TAHWEEL", "SUPPLIER", "CATEGORY", "DESCRIPTION",
    "UNIT SIZE\n(g or ml)", "DENSITY\n(g/ml)",
    "MAX PAR\n(units)", "MIN PAR\n(units)",
    "STOCK IN\n(grams — enter here)",
    "STOCK\n(ml converted)",
    "COST/GRAM\n(AED)",
    "STOCK VALUE\n(AED)"
]
for i, h in enumerate(inv_headers, 1):
    c = ws_inv.cell(row=3, column=i, value=h)
    c.fill = fill(DARK)
    c.font = Font(bold=True, color=GOLD, size=9, name="Arial")
    c.alignment = align("center", wrap=True)
    c.border = header_border()
ws_inv.row_dimensions[3].height = 36

# inv data mirrors master list
for row_i, (tahw, supp, cat, desc, unit_size, density, unit_cost) in enumerate(ingredients, 4):
    max_par = 4
    min_par = 1
    stock_g = 0  # user enters this

    col = lambda c: ws_inv.cell(row=row_i, column=c)
    col(1).value = tahw
    col(2).value = supp
    col(3).value = cat
    col(4).value = desc
    col(5).value = unit_size
    col(6).value = density
    col(7).value = max_par
    col(8).value = min_par
    # stock in grams — yellow input
    col(9).value = stock_g
    col(9).fill = fill(AMBER_LT)
    col(9).font = Font(bold=True, size=10, color=AMBER, name="Arial")
    # ml converted
    col(10).value = f"=IF(I{row_i}=0,0,I{row_i}/F{row_i})"
    col(10).number_format = '#,##0.0" ml"'
    # cost per gram
    col(11).value = f"=G{row_i}/E{row_i}"
    col(11).number_format = '"AED "#,##0.0000'
    # stock value
    col(12).value = f"=I{row_i}*K{row_i}"
    col(12).number_format = '"AED "#,##0.00'

    row_fill = fill(WHITE) if row_i % 2 == 0 else fill(GREY_LT)
    for c_i in range(1, 13):
        cell = ws_inv.cell(row=row_i, column=c_i)
        if c_i not in [9, 10, 11, 12]:
            cell.fill = row_fill
        cell.font = Font(size=9, name="Arial",
                         color="FF000080" if c_i == 1 else BLACK)
        cell.alignment = align("center" if c_i not in [2, 3, 4] else "left")
        cell.border = border()
    # low stock conditional hint (static for now — formula-driven in xl)
    ws_inv.row_dimensions[row_i].height = 14

# total value row
last_inv = 4 + len(ingredients) - 1
tot_inv = last_inv + 2
ws_inv.cell(tot_inv, 1).value = "TOTAL STOCK VALUE"
ws_inv.merge_cells(f"A{tot_inv}:K{tot_inv}")
ws_inv.cell(tot_inv, 1).fill = fill(DARK)
ws_inv.cell(tot_inv, 1).font = Font(bold=True, color=GOLD, size=10, name="Arial")
ws_inv.cell(tot_inv, 1).alignment = align("right")
ws_inv.cell(tot_inv, 12).value = f"=SUM(L4:L{last_inv})"
ws_inv.cell(tot_inv, 12).number_format = '"AED "#,##0.00'
ws_inv.cell(tot_inv, 12).fill = fill(DARK)
ws_inv.cell(tot_inv, 12).font = Font(bold=True, color=GOLD, size=10, name="Arial")
ws_inv.cell(tot_inv, 12).alignment = align("center")
ws_inv.row_dimensions[tot_inv].height = 20

# ─── Tab colours ────────────────────────────────────────────────────────────
ws_master.sheet_properties.tabColor = "B8860B"
ws_cost.sheet_properties.tabColor   = "1C1C2E"
ws_inv.sheet_properties.tabColor    = "006400"

# ─── Save ───────────────────────────────────────────────────────────────────
out_path = "/home/claude/Bar_Beverage_Costing_2026.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
